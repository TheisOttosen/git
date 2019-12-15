# imports.
import requests
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36 OPR/63.0.3368.56786"}

dataark = pd.read_excel("p-e.xlsm") # læser datasæt.
# print (dataark) # printer hele datasættet ud.
webadresser = dataark.adresser # laver en variabel med adresser.
# print (webadresser) # printer hele kolonnen med adresser ud.
# print (webadresser[0]) #  printer første adresse.

def getdata(index, url):
    page = requests.get(url, headers=headers) # skaffer siden
    soup = BeautifulSoup(page.content, "html.parser") # læser siden

    # finder data
    index = index
    navn = soup.find('h1', {"class":"D(ib) Fz(18px)"}).get_text() # finder aktienavn
    pris = soup.find('span', {"class":"Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)"}).get_text() # finder aktiepris
    data = soup.findAll('td', {"class": "Fz(s) Fw(500) Ta(end) Pstart(10px) Miw(60px)"}) # finder anden data
    p_e = data[2].get_text() # finder p-e
    p_bv = data[6].get_text() # finder p-bv
    kombineret = float(p_e) * float(p_bv) # finder kombineret

    #inputer data
    inputData(index, navn, pris, p_e, p_bv, kombineret)

def inputData(index, navn, pris, p_e, p_bv, kombineret):
    # laver index om til row
    index = index + 2
    
    # åbner dokument
    srcfile = openpyxl.load_workbook('p-e.xlsm',read_only=False, keep_vba= True)
    sheetname = srcfile.get_sheet_by_name('p-e')
    
    # skriver navn ind
    sheetname.cell(row=index,column=2).value = navn

    # skriver pris ind
    sheetname.cell(row=index,column=3).value = pris

    # skriver p-e ind 
    sheetname.cell(row=index,column=4).value = p_e
    
    # skriver p-bv ind
    sheetname.cell(row=index,column=5).value = p_bv

    # skriver kombineret ind
    sheetname.cell(row=index,column=6).value = kombineret

    # gemmer fil
    srcfile.save('p-e.xlsm')
    pass

for i in range(len(webadresser)):
    getdata(i, webadresser[i])